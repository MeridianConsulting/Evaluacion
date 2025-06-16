import pandas as pd
from openpyxl import load_workbook
import os
import unicodedata
import string

def normalizar_texto(texto):
    texto = str(texto).strip()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    return " ".join(texto.lower().split())

def limpiar_puntuacion(texto):
    return texto.translate(str.maketrans('', '', string.punctuation))

def procesar_excel_cargo_funciones(ruta_archivo):
    if not os.path.exists(ruta_archivo):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_archivo}")
    sql_cargo = []
    sql_funciones = []
    tabla_cargo = "cargo"
    tabla_funciones = "funciones"
    campo_id_cargo = "id_cargo"
    campo_nombre_cargo = "nombre_cargo"
    campo_desc_cargo = "descripcion_cargo"
    campo_objetivo_cargo = "objetivo_cargo"
    campo_proceso_gestion = "proceso_gestion"
    campo_id_cargo_func = "id_cargo"
    campo_titulo_funcion = "titulo_funcion"
    campo_descripcion_func = "descripcion_funcion"
    campo_tipo_funcion = "tipo_funcion"
    campo_hoja_funciones = "hoja_funciones"
    campo_fecha_creacion = "fecha_creacion"
    campo_fecha_actualiz = "fecha_actualizacion"
    campo_estado = "estado"
    default_desc_cargo = ""
    default_objetivo_cargo = ""
    default_proceso_gestion = ""
    default_tipo_funcion = "Específica"
    default_estado = "ACTIVO"
    cargos_existentes = set()
    next_id_cargo = 110
    wb = load_workbook(ruta_archivo, read_only=True)
    print("Iniciando procesamiento del archivo Excel...")
    for sheet_name in wb.sheetnames:
        print(f"Procesando hoja: {sheet_name}")
        try:
            df = pd.read_excel(ruta_archivo, sheet_name=sheet_name, header=None, engine='openpyxl')
            df = df.fillna('')
            header_row = None
            for i in range(len(df)):
                cell_value = str(df.iloc[i, 0]).strip()
                if not cell_value:
                    continue
                norm_cell = limpiar_puntuacion(normalizar_texto(cell_value))
                if ("funciones" in norm_cell and "profesional" in norm_cell) or \
                   ("funciones" in norm_cell and "especificas" in norm_cell and "del" in norm_cell and "cargo" in norm_cell):
                    header_row = i
                    print(f"Encabezado detectado en fila {i}")
                    break
            if header_row is None:
                print(f"[{sheet_name}] No se encontró encabezado válido. Se omite esta hoja.")
                continue
            # Usar el nombre de la hoja como el cargo y para el campo hoja_funciones
            cargo_str_final = sheet_name.strip().title()
            print(f"Cargo asignado (de la hoja): {cargo_str_final}")
            cargo_safe = cargo_str_final.replace("'", "''")
            norm_cargo = normalizar_texto(cargo_str_final)
            if norm_cargo in cargos_existentes:
                upsert_cargo = (f"UPDATE {tabla_cargo} SET {campo_desc_cargo}='{default_desc_cargo}', "
                                f"{campo_objetivo_cargo}='{default_objetivo_cargo}', {campo_proceso_gestion}='{default_proceso_gestion}' "
                                f"WHERE {campo_nombre_cargo}='{cargo_safe}';")
                print(f"Generando UPDATE para el cargo: {cargo_str_final}")
            else:
                upsert_cargo = (f"INSERT INTO {tabla_cargo} ({campo_id_cargo}, {campo_nombre_cargo}, {campo_desc_cargo}, "
                                f"{campo_objetivo_cargo}, {campo_proceso_gestion}) VALUES ({next_id_cargo}, '{cargo_safe}', "
                                f"'{default_desc_cargo}', '{default_objetivo_cargo}', '{default_proceso_gestion}');")
                cargos_existentes.add(norm_cargo)
                print(f"Generando INSERT para el cargo: {cargo_str_final}")
                next_id_cargo += 1
            sql_cargo.append(upsert_cargo)
            funciones_list = []
            for j in range(header_row + 1, len(df)):
                funcion_text = str(df.iloc[j, 0]).strip()
                if not funcion_text:
                    continue
                funciones_list.append(funcion_text)
            if not funciones_list:
                print(f"[{sheet_name}] No se encontraron funciones después del encabezado.")
                continue
            print(f"Se encontraron {len(funciones_list)} funciones en la hoja {sheet_name}.")
            # Usar un contador para generar una clave única para cada función
            function_counter = 1
            for funcion in funciones_list:
                funcion_safe = funcion.replace("'", "''")
                # El título de la función lo fijamos como "Función" (opcional) 
                # y movemos el texto extraído a descripcion_funcion
                titulo_funcion = "Función"
                descripcion_funcion = funcion_safe
                # Para la columna hoja_funciones, generamos una clave única combinando el nombre de la hoja y un contador
                hoja_funciones_val = cargo_safe + "_" + str(function_counter)
                function_counter += 1
                upsert_funcion = (f"INSERT INTO {tabla_funciones} ({campo_id_cargo_func}, {campo_titulo_funcion}, {campo_descripcion_func}, "
                                  f"{campo_tipo_funcion}, {campo_hoja_funciones}, {campo_fecha_creacion}, {campo_fecha_actualiz}, {campo_estado}) "
                                  f"SELECT {tabla_cargo}.{campo_id_cargo}, '{titulo_funcion}', '{descripcion_funcion}', '{default_tipo_funcion}', "
                                  f"'{hoja_funciones_val}', NOW(), NOW(), '{default_estado}' FROM {tabla_cargo} "
                                  f"WHERE {campo_nombre_cargo}='{cargo_safe}';")
                sql_funciones.append(upsert_funcion)
        except Exception as e:
            print(f"Error en la hoja [{sheet_name}]: {str(e)}")
            continue
    print("Procesamiento completado.")
    return sql_cargo, sql_funciones

if __name__ == "__main__":
    try:
        nombre_archivo = "GH-F-15_Evaluacion_desempeño 2024_v25.xlsm"
        script_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(script_dir, nombre_archivo)
        print("=" * 50)
        print(f"Directorio del script: {script_dir}")
        print(f"Ruta construida del Excel: {excel_path}")
        print(f"¿Existe el archivo?: {os.path.exists(excel_path)}")
        print("=" * 50)
        cargo_sql, funciones_sql = procesar_excel_cargo_funciones(excel_path)
        output_dir = os.path.join(script_dir, "sql_output")
        os.makedirs(output_dir, exist_ok=True)
        cargo_con_transaccion = "START TRANSACTION;\n" + "\n".join(cargo_sql) + "\nCOMMIT;\n"
        with open(os.path.join(output_dir, 'cargo.sql'), 'w', encoding='utf-8') as f:
            f.write(cargo_con_transaccion)
        funciones_con_transaccion = "START TRANSACTION;\n" + "\n".join(funciones_sql) + "\nCOMMIT;\n"
        with open(os.path.join(output_dir, 'funciones.sql'), 'w', encoding='utf-8') as f:
            f.write(funciones_con_transaccion)
        print("\nProceso completado. Archivos SQL generados en:")
        print(f"- {os.path.join(output_dir, 'cargo.sql')}")
        print(f"- {os.path.join(output_dir, 'funciones.sql')}")
        print(f"\nEstadísticas: Cargos: {len(cargo_sql)} sentencias, Funciones: {len(funciones_sql)} sentencias")
    except Exception as e:
        print(f"\nERROR: {str(e)}")
        print("Solución sugerida:")
        print("- Verifique que el archivo Excel esté en la misma carpeta que el script.")
        print("- Revise mayúsculas, espacios, puntuación y acentos en el contenido.")
        print("- Asegúrese de no tener el archivo abierto en otro programa.")
