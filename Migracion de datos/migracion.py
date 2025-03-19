import pandas as pd
from openpyxl import load_workbook
import os
import unicodedata

def normalizar_texto(texto):
    """
    Convierte el texto a minúsculas, elimina tildes y espacios extra.
    Ejemplo: 'Funciones_Profesional' -> 'funciones_profesional'
    """
    texto = str(texto)
    texto = ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )
    return texto.lower().strip()

def procesar_excel(ruta_archivo):
    """
    Procesa el archivo Excel y genera sentencias SQL para insertar/actualizar
    datos en la base de datos 'evaluacion', respetando la integridad referencial.
    
    Se generan sentencias para:
      - Tabla cargo: Insert/Update de cargos (usando nombre_cargo como clave única).
      - Tabla empleados: Insert/Update de empleados.
          * NOTA: En la tabla empleados el campo que referencia al cargo es "cargo" (VARCHAR)
            y es una FK que apunta a cargo.nombre_cargo.
      - Tabla funciones: Insert/Update de funciones, relacionando con cargo mediante id_cargo.
    
    Se buscan dos posibles encabezados:
      - "Funciones_Profesional" y "Funciones específicas del cargo"
    Se extrae la información a partir del primer encabezado detectado (insensible a mayúsculas/minúsculas),
    evitando duplicados y omitiendo el resto de la hoja.
    
    Importante: Se recomienda contar con un índice único en (id_cargo, titulo_funcion) en la tabla 'funciones'
    para evitar duplicados.
    """
    if not os.path.exists(ruta_archivo):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta_archivo}")
    
    # Listas para almacenar las sentencias SQL
    sql_empleados = []
    sql_funciones = []
    
    # Nombres de tablas
    tabla_cargo     = "cargo"
    tabla_empleados = "empleados"
    tabla_funciones = "funciones"
    
    # Campos de la tabla cargo
    campo_nombre_cargo = "nombre_cargo"
    campo_desc_cargo   = "descripcion_cargo"  # opcional
    
    # Campos de la tabla empleados
    campo_cedula            = "cedula"
    campo_nombre_empleado   = "nombre"
    campo_cargo_empleado    = "cargo"  # FK que referencia cargo.nombre_cargo
    campo_numero_telefonico = "numero_telefonico"
    campo_email             = "email"
    campo_compania          = "compania"
    campo_telefono_empresa  = "telefono_empresa"
    campo_telefono_internacional = "telefono_internacional"
    campo_contrasena        = "contrasena"
    
    # Campos de la tabla funciones
    campo_id_cargo_func    = "id_cargo"  # FK a cargo.id_cargo
    campo_titulo_funcion   = "titulo_funcion"
    campo_descripcion_func = "descripcion_funcion"
    campo_tipo_funcion     = "tipo_funcion"
    campo_fecha_creacion   = "fecha_creacion"
    campo_fecha_actualiz   = "fecha_actualizacion"
    campo_estado           = "estado"
    
    # Valores por defecto para empleados (modifica según tus necesidades)
    default_telefono    = "000000"
    default_email       = "desconocido@example.com"
    default_compania    = "Desconocida"
    default_tel_empresa = "000000"
    default_tel_internac = "N/A"
    default_contrasena  = "123456"
    
    # Cargos ya existentes en la tabla "cargo" (almacenados en forma normalizada)
    cargos_existentes = {
        normalizar_texto("Soporte Operativo Tipo 3A"),
        normalizar_texto("Asesor Administrativo"),
        normalizar_texto("Supervisor de Operaciones en Pozos Tipo 3"),
        normalizar_texto("Supervisor de Operaciones en Pozos Tipo 2"),
        normalizar_texto("Soporte Operativo Tipo 3B"),
        normalizar_texto("Profesional Senior para la ejecucion de actividades de la ODS No. 9770807 del contrato Matriz No. 30"),
        normalizar_texto("Director de Proyecto Ecopetrol"),
        normalizar_texto("Profesional Junior para la ejecucion de actividades de la ODS No. 9532986 del contrato Matriz No. 30"),
        normalizar_texto("Soporte Operativo Tipo 4A"),
        normalizar_texto("Profesional Junior para la ejecucion de actividades de la ODS No. 9770807 del contrato Matriz No. 30"),
        normalizar_texto("Supervisor de Operaciones en Pozos Tipo 1"),
        normalizar_texto("Soporte Operativo Tipo 2"),
        normalizar_texto("Subgerente"),
        normalizar_texto("Gerente General"),
        normalizar_texto("Analista de Gestion Humana y Nomina"),
        normalizar_texto("Asistente de Logistica"),
        normalizar_texto("Servicio Especializado Tipo 2 - Integridad"),
        normalizar_texto("Coordinador Contable y Financiero"),
        normalizar_texto("Gerente Administrativa Y Financiera"),
        normalizar_texto("Servicios Generales"),
        normalizar_texto("Aprendiz Etapa Lectiva"),
        normalizar_texto("Practicante Universitario"),
        normalizar_texto("Analista Contable"),
        normalizar_texto("Profesional de Proyectos"),
        normalizar_texto("Soporte Hseq"),
        normalizar_texto("Soporte Hseq II"),
        normalizar_texto("Programador Aprendiz SENA"),
        normalizar_texto("Soporte Operativo Tipo 5A"),
        normalizar_texto("Asistente administrativa y de gestion humana"),
        normalizar_texto("Asistente Administrativo"),
        normalizar_texto("Asistente de Nomina y gestion humana"),
        normalizar_texto("Asistente Contable"),
        normalizar_texto("Coordinador de Gestion Humana"),
        normalizar_texto("Tecnico Asistente Administrativa"),
        normalizar_texto("Asistente de Gestion Humana y Nomina"),
        normalizar_texto("Ingeniero(a) Asistente de Company Man para operaciones de Perforacion Completamiento y Workover  D1"),
        normalizar_texto("Ingeniero(a) Asistente de Company Man para operaciones de Perforacion Completamiento y Workover  D3"),
        normalizar_texto("Profesional B sico para la ejecucion de actividades de la ODS No. 9532986 del contrato Matriz No. 30"),
        normalizar_texto("Profesional de proyectos M1"),
        normalizar_texto("Profesional Senior para la ejecucion de actividades de la ODS No. 9532986 del contrato Matriz No. 30"),
        normalizar_texto("Ingeniero(a) Asistente de Company Man para operaciones de Perforacion Completamiento y Workover D2"),
        normalizar_texto("Profesional Junior para la ejecucion de actividades de la ODS No.9532986 del contrato Matriz No.303"),
        normalizar_texto("Profesional Junior para la ejecucion de actividades de la ODS No. 9814358 del contrato Matriz No. 30"),
        normalizar_texto("Coordinador Hseq"),
        normalizar_texto("Profesional Administrativa y de Gestion Humana, Proyectos"),
        normalizar_texto("Profesional Soporte en Campo"),
        normalizar_texto("Servicio Especializado Tipo 3A"),
        normalizar_texto("valor exacto del cargo"),
        normalizar_texto("Supervisor de Operaciones en Pozos Tipo 4"),
        normalizar_texto("Soporte IT Primer Nivel"),
        normalizar_texto("Contador Junior"),
        normalizar_texto("Aprendiz Etapa Practica"),
        normalizar_texto("Profesional Junior para la ejecucion de actividades de la ODS No.9532986 del contrato Matriz No. 303"),
        normalizar_texto("Soporte Hseq Proyectos")
    }
    
    # Establecemos el próximo id disponible (según los registros existentes, el máximo es 55)
    next_id = 56
    
    try:
        wb = load_workbook(ruta_archivo, read_only=True)
        
        for sheet_name in wb.sheetnames:
            try:
                df = pd.read_excel(ruta_archivo, sheet_name=sheet_name, header=None, engine='openpyxl')
                df = df.fillna('')
                
                # 1) Buscar el primer encabezado disponible: "Funciones_Profesional" o "Funciones específicas del cargo"
                header_row = None
                header_type = None
                for i in range(len(df)):
                    cell_value = df.iloc[i, 0]
                    norm_cell = normalizar_texto(cell_value)
                    if norm_cell.startswith("funciones_profesional"):
                        header_row = i
                        header_type = "funciones_profesional"
                        break
                    elif norm_cell.startswith("funciones especificas del cargo"):
                        header_row = i
                        header_type = "funciones_especificas"
                        break
                
                if header_row is None:
                    print(f"[{sheet_name}] No se encontró ningún encabezado válido ('Funciones_Profesional' o 'Funciones específicas del cargo'). Se omite esta hoja.")
                    continue
                
                # 2) Extraer cargo y nombre del empleado (comparación insensible a mayúsculas)
                cell_text = str(df.iloc[header_row, 0]).strip()
                if header_type == "funciones_profesional":
                    prefix = "Funciones_Profesional"
                else:
                    prefix = "Funciones específicas del cargo"
                
                cell_text_lower = cell_text.lower()
                prefix_lower = prefix.lower()
                idx = cell_text_lower.find(prefix_lower)
                if idx != -1:
                    cargo_nombre = cell_text[idx + len(prefix):].strip()
                else:
                    cargo_nombre = cell_text
                
                parts = cargo_nombre.split()
                if len(parts) < 2:
                    print(f"[{sheet_name}] No se pudo extraer el cargo y el nombre correctamente.")
                    continue
                
                cargo_extraido = parts[0]
                nombre_empleado_extraido = " ".join(parts[1:])
                
                # 3) Generar sentencia para la tabla cargo:
                # Se verifica si el cargo ya existe (comparando su forma normalizada).
                normalized_cargo = normalizar_texto(cargo_extraido)
                cargo_safe = cargo_extraido.replace("'", "''")
                if normalized_cargo in cargos_existentes:
                    # Si ya existe, generamos un UPDATE (actualizando la descripción a vacía)
                    upsert_cargo = (
                        f"UPDATE {tabla_cargo} SET {campo_desc_cargo} = '' WHERE {campo_nombre_cargo} = '{cargo_safe}';"
                    )
                else:
                    # Si no existe, generamos un INSERT asignando manualmente el id (con next_id)
                    upsert_cargo = (
                        f"INSERT INTO {tabla_cargo} (id_cargo, {campo_nombre_cargo}, {campo_desc_cargo}) "
                        f"VALUES ({next_id}, '{cargo_safe}', '');"
                    )
                    cargos_existentes.add(normalized_cargo)
                    next_id += 1
                
                # 4) Generar sentencia para la tabla empleados
                cedula_safe = sheet_name.replace("'", "''")
                nombre_safe = nombre_empleado_extraido.replace("'", "''")
                upsert_empleado = (
                    f"INSERT INTO {tabla_empleados} "
                    f"({campo_cedula}, {campo_nombre_empleado}, {campo_cargo_empleado}, "
                    f"{campo_numero_telefonico}, {campo_email}, {campo_compania}, "
                    f"{campo_telefono_empresa}, {campo_telefono_internacional}, {campo_contrasena}) "
                    f"SELECT "
                    f"'{cedula_safe}', "
                    f"'{nombre_safe}', "
                    f"'{cargo_safe}', "
                    f"'{default_telefono}', "
                    f"'{default_email}', "
                    f"'{default_compania}', "
                    f"'{default_tel_empresa}', "
                    f"'{default_tel_internac}', "
                    f"'{default_contrasena}' "
                    f"FROM dual "
                    f"ON DUPLICATE KEY UPDATE "
                    f"{campo_nombre_empleado} = VALUES({campo_nombre_empleado}), "
                    f"{campo_cargo_empleado} = VALUES({campo_cargo_empleado}), "
                    f"{campo_numero_telefonico} = VALUES({campo_numero_telefonico}), "
                    f"{campo_email} = VALUES({campo_email}), "
                    f"{campo_compania} = VALUES({campo_compania}), "
                    f"{campo_telefono_empresa} = VALUES({campo_telefono_empresa}), "
                    f"{campo_telefono_internacional} = VALUES({campo_telefono_internacional}), "
                    f"{campo_contrasena} = VALUES({campo_contrasena});"
                )
                
                sql_empleados.append(upsert_cargo)
                sql_empleados.append(upsert_empleado)
                
                # 5) Extraer las funciones a partir de la fila siguiente al encabezado detectado.
                # Se omiten las filas vacías para continuar la extracción.
                funciones_list = []
                for j in range(header_row + 1, len(df)):
                    funcion_text = str(df.iloc[j, 0]).strip()
                    norm_funcion = normalizar_texto(funcion_text)
                    if norm_funcion.startswith("funciones_profesional") or norm_funcion.startswith("funciones especificas del cargo"):
                        break
                    if not funcion_text:
                        continue
                    funciones_list.append(funcion_text)
                
                if not funciones_list:
                    print(f"[{sheet_name}] No se encontraron funciones para extraer después del encabezado.")
                    continue
                
                # 6) Para cada función, generar la sentencia de upsert en la tabla funciones.
                for funcion in funciones_list:
                    funcion_safe = funcion.replace("'", "''")
                    titulo_funcion = funcion_safe
                    upsert_funcion = (
                        f"INSERT INTO {tabla_funciones} "
                        f"({campo_id_cargo_func}, {campo_titulo_funcion}, {campo_descripcion_func}, "
                        f"{campo_tipo_funcion}, {campo_fecha_creacion}, {campo_fecha_actualiz}, {campo_estado}) "
                        f"SELECT {tabla_cargo}.id_cargo, '{titulo_funcion}', '', 'Específica', NOW(), NOW(), 'ACTIVO' "
                        f"FROM {tabla_cargo} "
                        f"WHERE {campo_nombre_cargo} = '{cargo_safe}' "
                        f"ON DUPLICATE KEY UPDATE "
                        f"{campo_descripcion_func} = VALUES({campo_descripcion_func}), "
                        f"{campo_tipo_funcion} = VALUES({campo_tipo_funcion}), "
                        f"{campo_fecha_actualiz} = NOW(), "
                        f"{campo_estado} = VALUES({campo_estado});"
                    )
                    sql_funciones.append(upsert_funcion)
            
            except Exception as e:
                print(f"Error en la hoja [{sheet_name}]: {str(e)}")
                continue
        
        return sql_empleados, sql_funciones
    
    except Exception as e:
        print(f"Error general al procesar el archivo: {str(e)}")
        raise

if __name__ == "__main__":
    try:
        # Ajusta el nombre de tu archivo Excel
        nombre_archivo = "GH-F-15_Evaluacion_desempeño 2024_v25.xlsm"
        
        script_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(script_dir, nombre_archivo)
        
        print("=" * 50)
        print(f"Directorio del script: {script_dir}")
        print(f"Ruta construida del Excel: {excel_path}")
        print(f"¿Existe el archivo?: {os.path.exists(excel_path)}")
        print("=" * 50)
        
        # Generamos las sentencias SQL
        empleados_sql, funciones_sql = procesar_excel(excel_path)
        
        # Directorio donde se guardarán los archivos SQL generados
        output_dir = os.path.join(script_dir, "sql_output")
        os.makedirs(output_dir, exist_ok=True)
        
        # Incluir transacciones para asegurar atomicidad
        empleados_con_transaccion = (
            "START TRANSACTION;\n"
            + "\n".join(empleados_sql)
            + "\nCOMMIT;\n"
        )
        with open(os.path.join(output_dir, 'empleados.sql'), 'w', encoding='utf-8') as f:
            f.write(empleados_con_transaccion)
        
        funciones_con_transaccion = (
            "START TRANSACTION;\n"
            + "\n".join(funciones_sql)
            + "\nCOMMIT;\n"
        )
        with open(os.path.join(output_dir, 'funciones.sql'), 'w', encoding='utf-8') as f:
            f.write(funciones_con_transaccion)
        
        print("\nProceso completado. Archivos SQL generados en:")
        print(f"- {os.path.join(output_dir, 'empleados.sql')}")
        print(f"- {os.path.join(output_dir, 'funciones.sql')}")
        print("\nEstadísticas:")
        print(f"- Empleados: {len(empleados_sql)} sentencias")
        print(f"- Funciones: {len(funciones_sql)} sentencias")
    
    except Exception as e:
        print(f"\nERROR: {str(e)}")
        print("Solución sugerida:")
        print("- Verifique que el archivo Excel esté en la misma carpeta que el script.")
        print("- Revise mayúsculas, espacios y acentos en el nombre del archivo.")
        print("- Asegúrese de no tener el archivo abierto en otro programa.")
